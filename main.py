import configparser as cfg
import datetime
import time
import sys
import os

import Get_copper as cop
import Get_Dollar as dol

import requests
import lxml
import openpyxl as xcl
from bs4 import BeautifulSoup
from selenium import webdriver


URL2 = "https://www.lme.com/en/Metals/Non-ferrous/LME-Copper#Trading+day+summary" # Цена на медь
URL1 = "https://cbr.ru/currency_base/daily/?UniDbQuery.Posted=True&UniDbQuery.To=" # Цена на доллар


def Get_Data(URL):
	files = os.listdir(".")
	excels = list(filter(lambda x: x.endswith(".xlsx"), files))
	excel = excels[0] 
	
	PATH = f"{os.getcwd()}\\{excel}"

	excel_file = xcl.load_workbook(PATH)

	sheets = excel_file.sheetnames[-1] # Выбрается последний (Массив страниц)
	employees_sheet = excel_file[sheets]

	row = datetime.datetime.today()
	_row = row.strftime("%d")
	
	if _row[0] == "0":
		row = _row[1] # 9 +7

	row = _row
	row = int(row) + 7

	date2 = datetime.date.today()
	date2 = date2 + datetime.timedelta(days = 1)
	date2 = str(date2)
	date2 = date2.replace("-", ".")	
	
	med = cop.parse_cooper()
	dollar = dol.parse_dollar(URL, date2)

	dollar = dollar.replace(",",".")	
	med = med.replace(",",".")
	
	employees_sheet.cell(row=row - 1, column=2).value = med ### Медь
	employees_sheet.cell(row=row, column=3).value = dollar ### Доллар

	excel_file.save(PATH)

Get_Data(URL1)

