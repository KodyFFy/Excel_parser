import configparser as cfg
import datetime
import time
import sys
import os
from tkinter.ttk import Progressbar 
import tkinter as tk
import tkinter.filedialog
from tkinter import messagebox

import Get_Dollar as dol

import requests
import lxml
import openpyxl as xcl
from bs4 import BeautifulSoup
from selenium import webdriver
config = cfg.ConfigParser()
config.read("cfg.ini")

print(config['Settings']['browser_id'])

def sts():
	URL1 = "https://cbr.ru/currency_base/daily/?UniDbQuery.Posted=True&UniDbQuery.To=" # Цена на доллар
	date2 = datetime.date.today()
	date2 = date2 + datetime.timedelta(days = 1)
	date2 = str(date2)
	date2 = date2.replace("-", ".")	

	dollar = dol.parse_dollar(URL1, date2)
	cooper = parse_cooper()

	dollar = dollar.replace(",",".")	


	PATH = config['Settings']['path']

	excel_file = xcl.load_workbook(PATH)

	sheets = excel_file.sheetnames[-1] # Выбрается последний (Массив страниц)
	employees_sheet = excel_file[sheets]

	row = datetime.datetime.today()
	_row = row.strftime("%d")
	
	if _row[0] == "0":
		row = _row[1] # 9 +7

	row = _row
	row = int(row) + 7
	employees_sheet.cell(row=row - 1, column=2).value = cooper ### Медь
	employees_sheet.cell(row=row, column=3).value = dollar ### Доллар

	excel_file.save(PATH)

	tkinter.messagebox.showinfo('Информация', "Запись в файл успешно завершена.")


def parse_cooper():
	config = cfg.ConfigParser()
	config.read("cfg.ini")
	binary = ["Drivers\\Chrome\\chromedriver.exe", "Drivers\\FireFox\\geckodriver.exe", "Drivers\\Opera\\operadriver.exe","Drivers\\Yandex\\yandexdriver.exe"]
	if config['Settings']['browser_id'] == "1":
		binar = binary[0]
		try:
			options = webdriver.ChromeOptions()
			options.add_argument('headless')
			driver = webdriver.Chrome(executable_path=binar, options=options)	
			driver.get("https://www.lme.com/en/Metals/Non-ferrous/LME-Copper#Trading+day+summary")
			print("Браузер Chrome успешно найден. Устанавливаю соединение...")
			print("Соединение с сайтом установлено успешно!")
		except:
			tkinter.messagebox.showerror('Ошибка', "Браузер Chrome не найден!")
			print("Браузер не найден или что-то пошло не так :(")
	elif config['Settings']['browser_id'] == "2":
		binar = binary[1]
		try:
			options = webdriver.FirefoxOptions()
			options.headless = True
			driver = webdriver.Firefox(executable_path=binar, options=options)	
			driver.get("https://www.lme.com/en/Metals/Non-ferrous/LME-Copper#Trading+day+summary")
			print("Браузер Firefox успешно найден. Устанавливаю соединение...")
			print("Соединение с сайтом установлено успешно!")
		except:
			tkinter.messagebox.showerror('Ошибка', "Браузер Firefox не найден!")
			print("Браузер не найден или что-то пошло не так :(")        
	elif config['Settings']['browser_id'] == "3":
		binar = binary[2]
		try:
			options = webdriver.Opera()
			options.add_argument('headless')
			driver = webdriver.Opera(executable_path=binar, options=options)	
			driver.get("https://www.lme.com/en/Metals/Non-ferrous/LME-Copper#Trading+day+summary")
			print("Браузер Opera успешно найден. Устанавливаю соединение...")
			print("Соединение с сайтом установлено успешно!")
		except:
			tkinter.messagebox.showerror('Ошибка', "Браузер Opera не найден!")
			print("Браузер не найден или что-то пошло не так :(")     
	elif config['Settings']['browser_id'] == "4":
		binar = binary[3]
		try:
			options = webdriver.ChromeOptions()
			options.add_argument('headless')
			driver = webdriver.Chrome(executable_path=binar, options=options)	
			driver.get("https://www.lme.com/en/Metals/Non-ferrous/LME-Copper#Trading+day+summary")
			print("Браузер Yandex успешно найден. Устанавливаю соединение...")
			print("Соединение с сайтом установлено успешно!")
		except:
			tkinter.messagebox.showerror('Ошибка', "Браузер Yandex не найден!")
			print("Браузер не найден или что-то пошло не так :(")     			
		driver.implicitly_wait(7)
		time.sleep(4)
		a = driver.find_elements_by_class_name("data-set-table__table")
		temp_mass = list()
		
		for elememt in a:
			temp_mass.append(elememt.text)
		
		mass = temp_mass[0].split()
		count = mass[4]
		
		b = len(count)

		if count[-1] == "0" and count[-2] == "0":
			count = count[:b - 3]
		else:
			pass
		return count


def sear():

	file_path = tk.filedialog.askopenfilename(filetypes = [("Excel файл(только .xlsx)", "*.xlsx")])
	PATH = file_path
	if file_path == "":
		file_path = config['Settings']['path']
	config['Settings']['path'] = file_path
	with open('cfg.ini', 'w') as cof:
		config.write(cof)

def change():
	if choice.get() == 1:
		config['Settings']['browser_name'] = "Google Chrome"
		config['Settings']['browser_id'] = "1"
	if choice.get() == 2:
		config['Settings']['browser_name'] = "Firefox"
		config['Settings']['browser_id'] = "2"
	if choice.get() == 3:
		config['Settings']['browser_name'] = "Opera"
		config['Settings']['browser_id'] = "3"
	if choice.get() == 4:
		config['Settings']['browser_name'] = "Yandex браузер"
		config['Settings']['browser_id'] = "4"
	with open('cfg.ini', 'w') as cof:
		config.write(cof)
def default_cfg():
		config['Settings']['browser_name'] = ""
		config['Settings']['browser_id'] = "0"
		config['Settings']['path'] = ""
		with open('cfg.ini', 'w') as cof:
			config.write(cof)

def open_cfg():
	os.startfile("cfg.ini")

def show_info():
	info_browse = config['Settings']['browser_name']
	paths = config['Settings']['path']

	mess1 = "Программа будет использовать браузер - " + info_browse
	mess2 = "Программа будет изменять файл по пути - " + paths

	tkinter.messagebox.showinfo('Браузер', mess1)
	tkinter.messagebox.showinfo('Путь', mess2)
def exitt():
	exit(0)

def how_to_use():
	os.startfile("how_to_use.txt")

### Конфиг
config = cfg.ConfigParser()
config.read("cfg.ini")

### Настройка окна
GUI = tk.Tk()
GUI.title("Parser to Excel")
GUI.geometry("460x210")
GUI.resizable(0,0)
GUI.iconbitmap("media\\ico.ico")

choice = tk.IntVar()
choice.set(int(config['Settings']['browser_id']))

menu = tk.Menu()
file_menu = tk.Menu(menu, tearoff=0)
file_menu.add_command(label="Сбросить конфиг программы", command = default_cfg)
file_menu.add_command(label="Изменить конфиг в ручную", command = open_cfg)
file_menu.add_separator()
file_menu.add_command(label="Как использовать программу?", command = how_to_use)
file_menu.add_separator()
file_menu.add_command(label="Закрыть программу", command = exitt)
menu.add_cascade(label="Настройки", menu=file_menu)
GUI.config(menu=menu)
### Выбор браузера

### Радио кнопки
browser1 = tk.Radiobutton(GUI, text = "Google Chrome", value = 1, variable = choice, command = change)
browser2 = tk.Radiobutton(GUI, text = "Firefox", value = 2, variable = choice, command = change)
browser3 = tk.Radiobutton(GUI, text = "Opera", value = 3, variable = choice, command = change)
browser4 = tk.Radiobutton(GUI, text = "Yandex браузер", value = 4, variable = choice, command = change)

### картинки
Chrome = tk.PhotoImage(file = "media\\1.png")
Firefox = tk.PhotoImage(file = "media\\2.png")
Opera = tk.PhotoImage(file = "media\\3.png")
Yandex_Browser = tk.PhotoImage(file = "media\\4.png")

lbl_Chrome = tk.Label(GUI,image = Chrome, width = 96, height = 96)
lbl_Chrome.grid(column = 0, row = 1)

lbl_Firefox = tk.Label(GUI,image = Firefox, width = 96, height = 96)
lbl_Firefox.grid(column = 1, row = 1) 

lbl_Opera = tk.Label(GUI,image = Opera, width = 96, height = 96)
lbl_Opera.grid(column = 2, row = 1) 

lbl_Yandex_Browser = tk.Label(GUI,image = Yandex_Browser, width = 96, height = 96)
lbl_Yandex_Browser.grid(column = 3, row = 1) 


### Путь
PATH = config['Settings']['path']

btn1 = tk.Button(GUI, text = "Выбрать файл",command = sear, font = (60))
btn1.grid(column = 1, row = 3)
btn2 = tk.Button(GUI, text = "Информация",command = show_info, font = (60))
btn2.grid(column = 2, row = 3)
btn3 = tk.Button(GUI, text = "Старт",command = sts ,font = (60))
btn3.place(x = 200, y = 157)



browser1.grid(column = 0, row = 2)
browser2.grid(column = 1, row = 2)
browser3.grid(column = 2, row = 2)
browser4.grid(column = 3, row = 2)

GUI.mainloop()
with open('cfg.ini', 'w') as cof:
	config.write(cof)